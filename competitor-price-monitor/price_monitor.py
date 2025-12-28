# =========================
# IMPORT REQUIRED LIBRARIES
# =========================

import requests                    # For making HTTP requests
import time                        # For adding delays between requests
import random                      # For random delay timing
from bs4 import BeautifulSoup      # For HTML parsing
from datetime import datetime      # For timestamps
import pandas as pd                # For data manipulation
import re                          # For cleaning price strings
import os                          # For file & directory handling
from openpyxl import load_workbook # For Excel post-processing
from openpyxl.styles import PatternFill  # For Excel cell coloring


# =========================
# STEP 1: CONFIGURATION
# =========================

# User-Agent header to mimic a real browser request
HEADER = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
}

# Central configuration for products & competitor URLs
CONFIG = {
    "products": [
        {
            "product_name": "A Light in the Attic",
            "competitors": [
                {
                    "name": "Amazon",
                    "url": "http://books.toscrape.com/catalogue/a-light-in-the-attic_1000/index.html"
                },
                {
                    "name": "Flipkart",
                    "url": "http://books.toscrape.com/catalogue/a-light-in-the-attic_1000/index.html"
                }
            ]
        }
    ]
}


# =========================
# STEP 2: SCRAPING FUNCTION
# =========================

def scrape_competitor(competitor):
    """
    Scrapes product information from a competitor website.
    Returns a dictionary containing product name, price,
    availability, competitor name, and timestamp.
    """

    try:
        # Random delay to avoid bot detection
        time.sleep(random.uniform(1, 3))

        # Send HTTP GET request
        response = requests.get(
            competitor["url"],
            headers=HEADER,
            timeout=10
        )

        # Handle failed requests
        if response.status_code != 200:
            print(f"ERROR: Could not access {competitor['name']}")
            return None

        response.raise_for_status()

        # Parse HTML content
        soup = BeautifulSoup(response.text, "html.parser")

        # NOTE: CSS selectors are site-specific
        product_name = soup.select_one("h1").get_text(strip=True)
        raw_price = soup.select_one(".price_color").get_text(strip=True)
        availability = soup.select_one(".instock").get_text(strip=True)

        return {
            "product_name": product_name,
            "competitor": competitor["name"],
            "raw_price": raw_price,
            "availability": availability,
            "timestamp": datetime.now()
        }

    except Exception as e:
        print(f"Error scraping {competitor['name']}: {e}")
        return None


# =========================
# STEP 3: DATA CLEANING
# =========================

def clean_price(raw_price):
    """
    Cleans raw price text and converts it to a float.
    Handles currency symbols and decimal formatting.
    """

    if pd.isna(raw_price):
        return None

    price_str = str(raw_price).strip()

    # Remove all non-numeric characters except . , -
    price_str = re.sub(r"[^\d.,-]", "", price_str)

    # Convert comma-based decimals to dot
    if price_str.count(",") == 1 and price_str.count(".") == 0:
        price_str = price_str.replace(",", ".")

    try:
        return float(price_str)

    except Exception:
        print(f"WARNING: Could not convert '{raw_price}' to float.")
        return None


# =========================
# STEP 4: PRICE ANALYSIS
# =========================

def calculate_price_changes(df):
    """
    Simulates previous prices and calculates price changes.
    Used for demonstration when historical data is unavailable.
    """

    previous_prices = []

    for i, price in enumerate(df["price"]):

        if price is None:
            previous_prices.append(None)

        # Simulate price drop
        elif i % 2 == 0:
            previous_prices.append(round(price * 0.90, 2))

        # Simulate price increase
        else:
            previous_prices.append(round(price * 1.10, 2))

    df["previous_price"] = previous_prices
    df["price_change"] = df["price"] - df["previous_price"]

    return df


# =========================
# STEP 5: EXCEL FORMATTING
# =========================

def format_excel(file_path):
    """
    Applies conditional formatting to Excel output:
    - Green for price decrease
    - Red for price increase
    """

    wb = load_workbook(file_path)
    ws = wb["price_change_summary"]

    green_fill = PatternFill(
        start_color="FFC6EFCE",
        end_color="FFC6EFCE",
        fill_type="solid"
    )

    red_fill = PatternFill(
        start_color="FFFFC7CE",
        end_color="FFFFC7CE",
        fill_type="solid"
    )

    price_change_col = 5  # Column E

    # Apply conditional formatting
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=price_change_col)

        if cell.value is not None:
            if cell.value > 0:
                cell.fill = red_fill
            elif cell.value < 0:
                cell.fill = green_fill

    # Auto-adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(file_path)


# =========================
# STEP 6: MAIN EXECUTION FUNCTION
# =========================

def main():
    """
    Main driver function:
    - Scrapes data
    - Cleans & analyzes prices
    - Exports formatted Excel report
    """

    print("Starting competitor price monitoring...")

    scraped_data = []

    for product in CONFIG["products"]:
        for competitor in product["competitors"]:
            try:
                data = scrape_competitor(competitor)
                if data:
                    scraped_data.append(data)
                print(f"Scraped data from {competitor['name']}")

            except Exception as e:
                print(f"Failed to scrape {competitor['name']}: {e}")

    if not scraped_data:
        print("No data scraped.")
        return

    # Convert to DataFrame
    df = pd.DataFrame(scraped_data)

    # Clean price column
    df["price"] = df["raw_price"].apply(clean_price)

    # Calculate price changes
    df = calculate_price_changes(df)

    # Create summary sheet
    summary_df = df[
        ["product_name", "competitor", "previous_price", "price", "price_change"]
    ]

    # Create report directory
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    base_dir = os.path.dirname(os.path.abspath(__file__))
    report_dir = os.path.join(base_dir, "report")
    os.makedirs(report_dir, exist_ok=True)

    # Excel file name
    filename = os.path.join(
        report_dir,
        f"price_summary_{timestamp}.xlsx"
    )

    # Write Excel report
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Raw_Data", index=False)
        summary_df.to_excel(
            writer,
            sheet_name="price_change_summary",
            index=False
        )

    # Apply formatting
    format_excel(filename)

    print(f"\n✅ Excel report generated successfully: {filename}")


# =========================
# STEP 7: SCRIPT ENTRY POINT
# =========================

if __name__ == "__main__":
    main()
